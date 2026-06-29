from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

BG_ROW    = "1E2333"
BG_ROWALT = "171B26"
BORDER_C  = "2A2F45"
BLUE_H    = "3B5BDB"
GREEN_H   = "2F9E44"
AMBER_H   = "E67700"
RED_H     = "C92A2A"
PURPLE_H  = "6741D9"
TEAL_H    = "0C8599"
GOLD_H    = "9A6700"
PINK_H    = "A61E4D"
WHITE     = "FFFFFF"
TEXT_DIM  = "8890A8"
TEXT_MID  = "C7D2FE"

STATUS_COLORS = {
    "Complete":       ("2F9E44", "FFFFFF"),
    "In Progress":    ("1864AB", "FFFFFF"),
    "Not Started":    ("C92A2A", "FFFFFF"),
    "Deprioritized":  ("495057", "AAAAAA"),
    "Needs GitHub":   ("E67700", "FFFFFF"),
    "R&D Phase":      ("6741D9", "FFFFFF"),
    "TBD":            ("495057", "CCCCCC"),
}

def score_color(score):
    if score == 0:   return "495057"
    if score < 30:   return "C92A2A"
    if score < 55:   return "E67700"
    if score < 75:   return "1864AB"
    return "2F9E44"

def thin_border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def row_fill(alt=False):
    return hfill(BG_ROWALT if alt else BG_ROW)

COLS = [
    ("Zone",        "zone",        9),
    ("Jira",        "jira",       10),
    ("Component",   "component",  24),
    ("Requirement / Deliverable", "requirement", 46),
    ("Score",       "score",       8),
    ("Status",      "status",     14),
    ("Due Date",    "due",        12),
    ("Notes for Validation", "notes", 40),
]

def make_sheet(name, accent, rows):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(row=1, column=1, value="  " + name + "  --  Sprint 01 / Lightgrid Energy")
    t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill = hfill(accent)
    t.alignment = Alignment(horizontal="left", vertical="center")
    t.border = thin_border()

    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
    s = ws.cell(row=2, column=1,
        value="  Validate with " + name.split()[0] + " - Jun 26 2026  |  Jira: lightgrid-energy.atlassian.net")
    s.font = Font(name="Arial", size=9, color=TEXT_DIM)
    s.fill = hfill("13161F")
    s.alignment = Alignment(horizontal="left", vertical="center")
    s.border = thin_border()

    ws.row_dimensions[3].height = 26
    for ci, (label, key, w) in enumerate(COLS, start=1):
        c = ws.cell(row=3, column=ci, value=label)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = hfill("1F2937")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(rows, start=4):
        ws.row_dimensions[ri].height = 44
        alt = (ri % 2 == 0)
        bf = row_fill(alt)

        for ci, (label, key, w) in enumerate(COLS, start=1):
            val = row.get(key, "")

            if key == "score":
                try:
                    sc = int(str(val).split("/")[0])
                except Exception:
                    sc = 0
                sc_hex = score_color(sc)
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
                c.fill = hfill(sc_hex)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()

            elif key == "status":
                bg, fg = STATUS_COLORS.get(val, ("374151", "CCCCCC"))
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=9, color=fg)
                c.fill = hfill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()

            elif key == "jira":
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=9, color="60A5FA", underline="single")
                c.fill = bf
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()

            elif key == "zone":
                zone_map = {
                    "Zone A": (BLUE_H, WHITE),
                    "Zone B": (GREEN_H, WHITE),
                    "Infra":  (AMBER_H, WHITE),
                    "--":     ("374151", TEXT_DIM),
                }
                bg, fg = zone_map.get(val, ("374151", TEXT_DIM))
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=9, color=fg)
                c.fill = hfill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()

            elif key == "component":
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=10, color=TEXT_MID)
                c.fill = bf
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border = thin_border()

            else:
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", size=9, color="D1D5DB")
                c.fill = bf
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                c.border = thin_border()

    ws.freeze_panes = "A4"

# ── MUJTABA ───────────────────────────────────────────────────
make_sheet("Mujtaba Khan", RED_H, [
    {"zone":"Zone A","jira":"SCRUM-38","component":"Lab Onboarding Checklist",
     "requirement":"Finalize checklist: data requirements, system access, customer qualification questions. Manish reviews for technical completeness.",
     "score":"40/100","status":"Not Started","due":"2026-06-15",
     "notes":"Overdue. Critical before university pilot. Draft first, then review with Manish."},
    {"zone":"Infra","jira":"SCRUM-36","component":"CI/CD Pipeline Ownership",
     "requirement":"Take ownership of CI/CD deployment strategy. GCP deploy approach, branch strategy, automated test gates.",
     "score":"28/100","status":"In Progress","due":"2026-06-28",
     "notes":"Mujtaba owns per Jun 25. Coordinate with Nishant who is pushing repos today/tomorrow."},
    {"zone":"Infra","jira":"SCRUM-39","component":"MLOps Pipeline (co-own w/ Nishant)",
     "requirement":"Model versioning, experiment tracking, dev to staging to prod promotion on GCP. Co-own with Nishant.",
     "score":"18/100","status":"Not Started","due":"2026-07-15",
     "notes":"Define ownership split with Nishant first before build starts."},
    {"zone":"Infra","jira":"SCRUM-31","component":"Zone A-B Interface Contract (sign-off)",
     "requirement":"Formal sign-off once Manish + Nishant complete spec. Review and lock. Hard blocker on E2E test.",
     "score":"10/100","status":"Not Started","due":"2026-06-28",
     "notes":"Manish drafts, Nishant co-signs from Zone B, Mujtaba final sign-off."},
    {"zone":"--","jira":"--","component":"Jira Team Invites (BLOCKING)",
     "requirement":"Invite Rakshit, Nishant, Manish, Tilak, Krish to lightgrid-energy.atlassian.net. No team member has a Jira account yet.",
     "score":"0/100","status":"Not Started","due":"2026-06-26",
     "notes":"BLOCKING everything. Must happen before sprint tracking can start. Send invites today."},
    {"zone":"--","jira":"--","component":"Krish Redeployment Decision",
     "requirement":"DVFS + TCS deprioritized. Decide where Krish contributes: (A) Eval gate testing, (B) Zone B protocols MQTT/BACnet, (C) 50MW sim data prep.",
     "score":"--","status":"TBD","due":"2026-06-26",
     "notes":"Recommend Option B - Zone B protocols. Nishant overloaded and Krish has hardware/systems background."},
])

# ── MANISH ────────────────────────────────────────────────────
make_sheet("Manish Chauhan", PINK_H, [
    {"zone":"Infra","jira":"SCRUM-31","component":"Zone A-B Interface Contract (draft)",
     "requirement":"Define full API contract: endpoints, payload schema, error handling, telemetry return format. Nishant must co-sign from Zone B side.",
     "score":"10/100","status":"Not Started","due":"2026-06-28",
     "notes":"CRITICAL - overdue Jun 12. Blocks E2E integration test (hard gate Jun 28). Priority #1 for Manish."},
    {"zone":"Zone A","jira":"SCRUM-33","component":"Decision Engine + Operator UI",
     "requirement":"Output format defined, basic rendering working. 3 audience views: Executive (risk profile, sign-off), Watcher/Facility Manager (config), Tenant (opt-in, savings).",
     "score":"32/100","status":"In Progress","due":"2026-07-04",
     "notes":"6 personas fully mapped per Granola. Wireframes in progress. RBAC groundwork with Nishant. Validate wireframe % tomorrow."},
    {"zone":"Zone A","jira":"SCRUM-34","component":"Tenant Surface",
     "requirement":"Flex-tier opt-in interface. Savings attribution display wired to AE output. Educational MVP demo for first pilots.",
     "score":"20/100","status":"Not Started","due":"2026-07-04",
     "notes":"Depends on AE output format being locked first. Manish owns design."},
    {"zone":"Zone A","jira":"SCRUM-13","component":"Accountability Engine (arch review)",
     "requirement":"Review and validate dual-brain architecture: checklist/soul layer + MPC brain. 11 rules, 4 clusters (Know/Decide/Limit/Give Back).",
     "score":"28/100","status":"In Progress","due":"2026-06-25",
     "notes":"Rakshit + Nishant building. Manish reviews architecture and signs off on design."},
    {"zone":"Zone B","jira":"--","component":"Auth / RBAC Groundwork",
     "requirement":"Role-based access control foundation for multi-tenant dashboard. Define role hierarchy: exec, operator, admin, tenant. Co-build with Nishant.",
     "score":"15/100","status":"In Progress","due":"2026-07-04",
     "notes":"Mentioned in Granola. Define role hierarchy before UI build starts."},
])

# ── RAKSHIT ───────────────────────────────────────────────────
make_sheet("Rakshit Mishra", PURPLE_H, [
    {"zone":"Zone A","jira":"SCRUM-16","component":"Geo-shift RL Optimizer",
     "requirement":"Train geo-based RL pathway, validate + pass all 8 eval gates. ITransformer architecture. Physics equations integrated.",
     "score":"42/100","status":"In Progress","due":"2026-06-25",
     "notes":"In progress. Policy-based, needs production upgrade. V1.5 training ongoing. Confirm gate pass status tomorrow."},
    {"zone":"Zone A","jira":"SCRUM-17","component":"Time-shift RL Optimizer",
     "requirement":"Train time-based RL pathway, validate against eval gates. 2-5 min ahead predictions, low latency. Outcomes: 31% lower peak, 27% lower carbon, 16% cheaper.",
     "score":"42/100","status":"In Progress","due":"2026-06-25",
     "notes":"Parallel to geo-shift. Same gate requirements. Confirm % complete tomorrow."},
    {"zone":"Zone A","jira":"SCRUM-11","component":"Classifier",
     "requirement":"Flex-tier 0-3, priority scoring, tenant sensitivity, only-what-is-given enforcement. Hierarchical: critical / medium / flexible tiers.",
     "score":"25/100","status":"R&D Phase","due":"2026-06-20",
     "notes":"Overdue. Tilak supporting. Move from R&D to build phase. Technical classification can be done now."},
    {"zone":"Zone A","jira":"SCRUM-12","component":"Rescheduler",
     "requirement":"Dispatch layer + trigger gates (grid signal, $/CO2 delta threshold, emergency). Trigger gate logic lives here - NOT in the optimizer.",
     "score":"32/100","status":"In Progress","due":"2026-06-25",
     "notes":"Rakshit owns (corrected from Manish). MPC checklist layer as governance wrapper. Confirm approach tomorrow."},
    {"zone":"Zone A","jira":"SCRUM-13","component":"Accountability Engine (build)",
     "requirement":"11 rules, 4 clusters. Dual-brain: soul/checklist layer always-on + MPC second brain 10-step lookahead. Co-build with Nishant.",
     "score":"28/100","status":"In Progress","due":"2026-06-25",
     "notes":"Interesting architecture + thesis per Mujtaba. Hasti will expand later. Validate thesis completeness tomorrow."},
    {"zone":"Zone A","jira":"SCRUM-27","component":"Eval Framework (lead - all 8 gates)",
     "requirement":"Own gates 0-8: data quality, ML quality, physics compliance ASHRAE A1, calibration, generalization, decision quality, RL policy quality, numerical accuracy.",
     "score":"50/100","status":"In Progress","due":"2026-06-23",
     "notes":"Rakshit main owner (corrected). V2 gates 1-4 built. ~2-3 hrs of tests remaining. Tilak supporting."},
    {"zone":"Zone A","jira":"SCRUM-30","component":"Load Classifications",
     "requirement":"10/50/100 MW load profiles + flexibility tiers + ramp rates. 4 types: AI Training 50%, Inference 25%, Batch Analytics 15%, Checkpointing 10%.",
     "score":"30/100","status":"In Progress","due":"2026-06-20",
     "notes":"Feeds directly into classifier. Benchmark sources: MLPerf, Alibaba cluster, BitBrains, Google Cluster."},
])

# ── NISHANT ───────────────────────────────────────────────────
make_sheet("Nishant Raghuvanshi", TEAL_H, [
    {"zone":"Zone A","jira":"SCRUM-26","component":"Data Ingestion - Push to GitHub",
     "requirement":"All 3 pipelines (Grid APIs/weather, Facility telemetry, Workload traces) complete and pushed to GitHub. Verified against schema contract.",
     "score":"72/100","status":"Needs GitHub","due":"2026-06-25",
     "notes":"Mujtaba confirms complete. Push target: today/tomorrow (Jun 25/26) per Granola. Validate commit is live."},
    {"zone":"Zone A","jira":"SCRUM-18","component":"AE Low-Fi Validator Loop",
     "requirement":"Dispatch -> AE validates -> recursive refinement until output passes -> passes to hi-fi DT. Threshold-based N runs.",
     "score":"35/100","status":"In Progress","due":"2026-07-02",
     "notes":"Nishant owns (corrected). GCP pipeline being built - checklist + MPC versions in parallel. Tilak framed internal sim for this."},
    {"zone":"Zone B","jira":"SCRUM-19","component":"Historian Connector - Deploy-Ready",
     "requirement":"Confirm deploy-ready, push to GitHub. BMS/EMS/DCIM nodes, ingestion to GCP. Humber tag list verified. MQTT scope confirmed.",
     "score":"80/100","status":"In Progress","due":"2026-06-14",
     "notes":"Pretty well done per Mujtaba. Overdue - confirm GitHub status. OPC UA first; MQTT scope after Humber meeting."},
    {"zone":"Zone B","jira":"SCRUM-20","component":"OPC UA Layer",
     "requirement":"Standard endpoints, master tag list integration, read/write validated end-to-end. Needs Humber to expose server URL, port, node structure.",
     "score":"65/100","status":"In Progress","due":"2026-06-24",
     "notes":"~70% per Mujtaba. TBD - confirm with team tomorrow. Coordinate Humber server access."},
    {"zone":"Zone B","jira":"SCRUM-21","component":"MQTT Plugin",
     "requirement":"Assess scope post-Humber, then implement + validate. Config-based framework template for DC types without DCIM.",
     "score":"15/100","status":"Not Started","due":"2026-06-24",
     "notes":"Scope TBD after Humber meeting. Config template approach for MQTT-based DCs."},
    {"zone":"Zone B","jira":"SCRUM-22","component":"BACnet Plugin",
     "requirement":"Assess scope + implement + validate. Part of multi-protocol historian framework.",
     "score":"15/100","status":"Not Started","due":"2026-06-24",
     "notes":"Same config-template approach as MQTT. Prioritize after OPC UA confirmed."},
    {"zone":"Zone B","jira":"SCRUM-23","component":"Modbus Plugin",
     "requirement":"Assess scope + implement + validate.",
     "score":"15/100","status":"Not Started","due":"2026-06-24",
     "notes":"Third protocol in stack. Scope post-Humber interaction."},
    {"zone":"Zone B","jira":"SCRUM-24","component":"GCP Historian Deploy",
     "requirement":"Deploy + validate end-to-end on GCP. Zone A dispatches flowing into GCP time-series storage.",
     "score":"65/100","status":"In Progress","due":"2026-06-20",
     "notes":"Overdue. Confirm deployment status tomorrow."},
    {"zone":"Zone B","jira":"SCRUM-25","component":"GCP Telemetry Return Path",
     "requirement":"Telemetry flowing from PLC/physical asset back to GCP. Latency measured and within budget.",
     "score":"20/100","status":"Not Started","due":"2026-06-25",
     "notes":"Blocked on Humber OPC UA connection first. Architecture designed - needs physical endpoint."},
    {"zone":"Infra","jira":"SCRUM-31","component":"Zone A-B Interface Contract (co-sign)",
     "requirement":"Review + co-sign API contract from Zone B perspective. Confirm payload schema and telemetry return format.",
     "score":"10/100","status":"Not Started","due":"2026-06-28",
     "notes":"Manish drafts, Nishant co-signs. Hard blocker on E2E test."},
    {"zone":"Infra","jira":"SCRUM-36","component":"CI/CD Pipeline (execution)",
     "requirement":"Push all remaining repos to GitHub. Repo structure, branch strategy per Mujtaba direction.",
     "score":"28/100","status":"In Progress","due":"2026-06-25",
     "notes":"Mujtaba owns strategy. Nishant executes. Repos target: today/tomorrow."},
    {"zone":"Infra","jira":"SCRUM-39","component":"MLOps Pipeline (co-own w/ Mujtaba)",
     "requirement":"Model versioning, experiment tracking, dev to staging to prod on GCP.",
     "score":"18/100","status":"Not Started","due":"2026-07-15",
     "notes":"Define ownership split with Mujtaba first."},
    {"zone":"Infra","jira":"SCRUM-35","component":"GCP Monthly Cost Estimate",
     "requirement":"Pull GCP monthly cost estimate, coordinate with Rakshit/Krish/Tilak on compute needs, send to Mujtaba.",
     "score":"0/100","status":"Not Started","due":"2026-06-12",
     "notes":"Overdue since Jun 12. Quick task - pull from GCP console and send today."},
    {"zone":"Infra","jira":"--","component":"Internal Simulation 50MW (co-own w/ Tilak)",
     "requirement":"GCP pipeline for 50MW reference DC simulation. ERCOT-based, physics engine, workload graphs, full-stack E2E.",
     "score":"32/100","status":"In Progress","due":"2026-07-25",
     "notes":"Co-own with Tilak. Tilak defines DC spec (task 1); Nishant builds GCP pipeline."},
])

# ── TILAK ─────────────────────────────────────────────────────
make_sheet("Tilak", GOLD_H, [
    {"zone":"Zone A","jira":"SCRUM-30","component":"50MW Reference DC - Task 1 (PRIMARY)",
     "requirement":"Define the 50MW reference DC simulation: solar/battery/asset mix, workload profiles, flexibility tiers, ramp rates. Foundation for internal simulator.",
     "score":"30/100","status":"In Progress","due":"2026-06-20",
     "notes":"PRIMARY task per Granola. Send task 1 doc to Nishant ASAP. Nishant + Rakshit supporting on spec sheet."},
    {"zone":"Zone A","jira":"SCRUM-26","component":"Data Schema Contract",
     "requirement":"Define fields, types, validation rules for all 3 ingestion sources. Needed to formally verify Nishant data ingestion pipelines.",
     "score":"40/100","status":"Not Started","due":"2026-06-14",
     "notes":"Overdue Jun 10. Coordinate with Nishant - ingestion complete but formal schema doc needed."},
    {"zone":"Zone A","jira":"SCRUM-27","component":"Eval Gates 1+2 (Rakshit leads, Tilak supports)",
     "requirement":"Data quality gate + physical out-of-bounds gate. These block the entire pipeline.",
     "score":"50/100","status":"In Progress","due":"2026-06-20",
     "notes":"~2-3 hrs of tests remaining. Rakshit leads. Gate 1+2 must pass before optimizer output reaches rescheduler."},
    {"zone":"Zone A","jira":"SCRUM-28","component":"Eval Gates 3+4 (support)",
     "requirement":"Overfit detection + per-variable accuracy gate.",
     "score":"45/100","status":"Not Started","due":"2026-06-23",
     "notes":"Depends on gates 1+2 passing first."},
    {"zone":"Zone A","jira":"SCRUM-29","component":"Eval Gates 5+6 (support)",
     "requirement":"Latency budget gate + zero physics violations (ASHRAE A1 - most critical).",
     "score":"40/100","status":"Not Started","due":"2026-06-23",
     "notes":"Physics compliance is the most important gate. All must pass before optimizer output is trusted."},
    {"zone":"Zone A","jira":"SCRUM-11","component":"Classifier (support - hierarchical jobs)",
     "requirement":"Support Rakshit on hierarchical job classification: critical/medium/flexible. 4 workload types (Training 50%/Inference 25%/Batch 15%/Checkpointing 10%).",
     "score":"25/100","status":"R&D Phase","due":"2026-06-20",
     "notes":"Overdue. Move from R&D to build. Technical part can be done now."},
    {"zone":"Infra","jira":"SCRUM-32","component":"E2E Integration Test - HARD GATE",
     "requirement":"Full pipeline: ingestion -> classifier -> eval -> opt engine -> rescheduler -> AE -> historian -> OPC UA -> telemetry return. ALL gates passing.",
     "score":"5/100","status":"Not Started","due":"2026-06-28",
     "notes":"HARD GATE. If this fails, lab deployment does not proceed. Nishant + Rakshit on-call to fix. Blocked on interface contract."},
    {"zone":"Infra","jira":"--","component":"Internal Simulation 50MW (co-own w/ Nishant)",
     "requirement":"Co-own with Nishant. Define reference DC specs (task 1) then validate full stack through simulation.",
     "score":"32/100","status":"In Progress","due":"2026-07-25",
     "notes":"Tilak defines DC spec; Nishant builds GCP pipeline. ERCOT + physics-based. All eval gates must pass."},
])

# ── KRISH ─────────────────────────────────────────────────────
make_sheet("Krish Patel", PINK_H, [
    {"zone":"Zone A","jira":"SCRUM-14","component":"DVFS - DEPRIORITIZED",
     "requirement":"Dynamic voltage + frequency scaling. 3 architectures proposed, option 3 selected. 100ms latency at current spec.",
     "score":"30/100","status":"Deprioritized","due":"--",
     "notes":"Blocked on multi-GPU (needs 2-3 NVLink GPUs). Single-GPU failed. Deprioritized for this sprint per Mujtaba Jun 25."},
    {"zone":"Zone A","jira":"SCRUM-15","component":"TCS Cooling - DEPRIORITIZED",
     "requirement":"Thermal control system + cooling optimization research.",
     "score":"20/100","status":"Deprioritized","due":"--",
     "notes":"Secondary for this sprint. Resume when GPU infrastructure available."},
    {"zone":"--","jira":"--","component":"Redeployment Decision - TBD",
     "requirement":"DVFS + cooling deprioritized. Options: (A) Eval gate testing support with Rakshit/Tilak, (B) MQTT/BACnet/Modbus protocol research for Zone B, (C) 50MW sim data prep.",
     "score":"--","status":"TBD","due":"2026-06-26",
     "notes":"DECIDE TOMORROW. Recommend Option B - Zone B protocols. Nishant overloaded, Krish has hardware/systems background."},
    {"zone":"Infra","jira":"SCRUM-35","component":"GCP Compute Estimate (input to Nishant)",
     "requirement":"Provide GPU compute requirements to Nishant for monthly GCP cost estimate based on DVFS work so far.",
     "score":"--","status":"TBD","due":"2026-06-26",
     "notes":"Nishant consolidates all inputs and sends to Mujtaba."},
])

out = r"C:\Users\mujta\OneDrive\Documents\Crypto\Robot Work\chief-of-staff\Sprint01_TeamRequirements.xlsx"
wb.save(out)
print("Saved:", out)
